# -*- coding: utf-8 -*-
# File  : ReadExcelClass.py
# Author: water
# Date  : 2019/9/12

import xlrd
import openpyxl


class ReadExcel:

    def __init__(self,excel_path):
        self.wb:workbook = xlrd.open_workbook(self.excel_path)
        self.ws_common = wb.sheet_by_name("common")
        self.ws_case = wb.sheet_by_name("case")
        self.datas = []

    def close_excel(self):
        pass

    def __get_case(self):
        self.case = dict()
        self.cases = list()
        max_row = self.ws_case.nrows
        for i in range(1,max_row):
            self.case['step'] = ws_case.cell(i,0).value
            self.case['test_desc'] = ws_case.cell(i, 0).value
            self.case['out_put'] = ws_case.cell(i, 0).value
            self.case['request_name'] = ws_case.cell(i, 0).value
            self.case['resquest_param'] = ws_case.cell(i, 0).value
            self.case['step'] = ws_case.cell(i, 0).value

    def __get_common(self):
        self.common = dict()
        max_row = self.ws_common.nrows
        for i in range(max_row):
            self.common[self.ws_common.cell(i,0).value] = self.ws_common.cell(i,1).value

    def get_data(self):
        pass



"""
    import openpyxl
    data = openpyxl.load_workbook('excel_test.xlsx')
    print(data.get_named_ranges())  # 输出工作页索引范围
    print(data.get_sheet_names())  # 输出所有工作页的名称
    # 取第一张表
    sheetnames = data.get_sheet_names()
    table = data.get_sheet_by_name(sheetnames[0])
    table = data.active
    print(table.title)  # 输出表名
    nrows = table.max_row  # 获得行数
    ncolumns = table.max_column  # 获得行数
    values = ['E', 'X', 'C', 'E', 'L']
    for value in values:
        table.cell(nrows + 1, 1).value = value
        nrows = nrows + 1
    data.save('excel_test.xlsx')

"""